"""
This code is prefect code for scrapping of metro.ca wesite
This code run well python3.82
If  you run this code in higher or lower version, there are some errors like undetected_chromedriver and distutils error

"""
import csv
import time
import logging
import undetected_chromedriver as webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook


logger = logging.getLogger('')
logger.setLevel(logging.DEBUG)

file_handler = logging.FileHandler('Electrical.log', encoding='utf-8')
formatter = logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s %(module)s - %(funcName)s: %(message)s',
                              datefmt='%Y-%m-%d %H:%M:%S')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

class electrical_scraper:
    
    
    def __init__(self):
        self.driver = None
        self.products = []
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
    def run_browser(self, url):
        service = Service(executable_path=r'D:/Browser App/chromedriver-win64/chromedriver.exe')
        
        chrome_options = webdriver.ChromeOptions()
        #chrome_options.add_argument("--headless")
        chrome_options.add_argument('--window-size=1920,1080')  # Set an appropriate window size
        #chrome_options.add_argument("--use_subprocess")
        #chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument('--no-sandbox')
        #chrome_options.add_argument('--disable-dev-shm-usage')
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.driver.maximize_window()
        self.driver.get(url)

    def scrape_website(self):
        url = "https://www.electrical.com/Products/Active-Harmonic-Filter"
        self.run_browser(url)
        '''
        # Initially locate the list of category elements
        categories_count = len(WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, './/li[contains(@class, "single-category")]'))))
        
        for index in range(categories_count):
            # Re-locate the list of categories and the specific category_li on each iteration
            category_lis = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, './/li[contains(@class, "single-category")]')))
            category_li = category_lis[index]
            category = category_li.find_element(By.XPATH, './/h3')
            category_name = category.text.replace('/', '_')
            print(category_name)
             # Create a new sheet for each category with its name
            if category_name in self.workbook.sheetnames:
                sheet = self.workbook[category_name]
            else:
                sheet = self.workbook.create_sheet(title=category_name)
            try:
                # Navigate to the category page
                a_element = category_li.find_element(By.XPATH, './/div[contains(@class, "category-image")]/a')
                a_element.click()
                subcategory_list = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, './/li[contains(@class, "single-category-small")]/h3'))) 
                # Wait for the h3 elements to be loaded in the new category page
                try:
                    WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, './/li[contains(@class, "single-category-small")]/h3')))
                    sub_categories = self.driver.find_elements(By.XPATH, './/li[contains(@class, "single-category-small")]/h3')
                    
                    for idx, sub_cat in enumerate(sub_categories, start=1):
                        sheet.append([sub_cat.text])  # Write each sub-category to the sheet

                    
                except :
                    print("Timed out waiting for h3 elements to load.")
                # Perform your scraping tasks on the category page here...
                # For example, scrape details of the category

                # Navigate back to the list of categories
                self.driver.back()

                # Add a wait here if necessary, to ensure the page has fully loaded
                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, './/li[contains(@class, "single-category")]')))

            except Exception as e:
                print(f"Failed to navigate and scrape category: {e}")
                self.driver.back()  # Ensure we navigate back even if an error occurs
            
            # Add a brief pause if necessary to avoid hitting the server too hard
        '''
        time.sleep(50)
    
    
    def save_workbook(self, filename):
        self.workbook.save(filename=filename)
if __name__ == "__main__":
    electrical_scraper = electrical_scraper()
    electrical_scraper.scrape_website()
    #electrical_scraper.save_workbook('electrical_categories.xlsx')